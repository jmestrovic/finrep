# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
import xlrd
import os
import fnmatch
import re
import sys
from django.core.wsgi import get_wsgi_application
import logging
import shutil
from datetime import datetime


input_subfolder = 'input'
archive_subfolder = 'archive'
error_subfolder = 'error'
input_pattern = '*.xls*'
django_project_path = '..\\web'
company_name_key = 'NAME'
company_abbreviation_key = 'ABBREV'
gfi_expected_items_no = 170     # expected No of positions in GFI
save_status_ok = 1
save_status_error = -1
gfi_orig_filename_key = 'GFI_ORIG_FILENAME'
gfi_dest_filename_key = 'GFI_DEST_FILENAME'
gfi_year_key = 'GFI_YEAR'

# prije promjene radnog foldera (radi Djanga) preuzimam
# folder u kojemu se nalazi skripta koja se pokreće
script_folder = os.path.dirname(os.path.realpath(__file__))
input_folder = os.path.join(script_folder, input_subfolder)
archive_folder = os.path.join(script_folder, archive_subfolder)
error_folder = os.path.join(script_folder, error_subfolder)

# logging configuration
logging.basicConfig(filename='finrep.log', level=logging.DEBUG)

# This is so Django knows where to find stuff.
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "finrep.settings")
sys.path.append(django_project_path)

# This is so my local_settings.py gets loaded.
os.chdir(django_project_path)

# This is so models get loaded.
application = get_wsgi_application()

# import models
from gfi.models import *


def get_all_input_files(folder, pattern):
    '''
    browse input_folder; search  files whose name matches with input_pattern
    '''
    for root, dirs, files in os.walk(folder):
        for basename in files:
            if fnmatch.fnmatch(basename, pattern):
                filename = os.path.join(root, basename)
                yield filename


def open_xls_as_xlsx(fname):
    '''
    load xls file with file name fname, convert it to
    xlsx format (in memory) an return converted file
    '''
    xlsBook = xlrd.open_workbook(filename=fname)
    workbook = Workbook()
    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

    return workbook


def is_row_empty(row):
    '''
    checks every cell value in array row
    returns True if all values are empty or None
    '''
    is_empty = True
    for cell in row:
        if str(cell.value):
            is_empty = False
            break
    return is_empty


def is_data_row(row):
    '''
    preskoči potpuno prazne retke te one koji u kolonama 1-6 imaju
    neki podatak (takvi reci ne sadrže financijske podatke)
    '''
    is_data_row = (not is_row_empty(row)) \
        and is_row_empty(row[1:7]) \
        and isinstance(row[8].value, (int, float))
    return is_data_row


def eligible_rows(rows):
    '''
    Preskače prvih x redaka na sheetu (zaglavlje)
    Također, preskače i sve ostale retke koji ne sadrže podatke
    '''
    r = []
    is_header = True
    for row in rows:
        if (str(row[0].value).strip().upper().startswith('A)  POTRAŽIVANJA') or(str(row[0].value).strip().startswith('I. POSLOVNI'))):
            is_header = False
        if (not is_header) and is_data_row(row):
            r.append(row)
    return r


def find_company_name(rows):
    '''
    Traži redak koji sadrži naziv tvrtke, te taj naziv ekstrahira i vraća
    '''
    name = ''
    pattern = '\AOBVEZNIK:*\s*(.*)[\s\t]*\Z'
    for row in rows:
        val = (row[0].value).replace('_', ' ').strip().upper()
        if val.startswith('OBVEZNIK'):
            try:
                res = re.search(pattern, val)
                name = res.groups(1)[0]
            except AttributeError:
                pass  # not found
            break
    return name


def find_company_abbreviation_and_gfi_year(filename):
    '''
    Extracts company abbreviation and year from filename
    '''
    _, filename = os.path.split(filename)
    abbrev = ''
    pattern = '\A(\w*)-fin(\d*)-.*\Z'
    try:
        res = re.search(pattern, filename)
        abbrev = res.groups(1)[0]
        year = res.groups(1)[1]
    except AttributeError:
        pass  # not found
    return (abbrev, year)


def check_filename(filename):
    '''
    Checks if filename is in line with expected
    '''
    _, filename = os.path.split(filename)
    pattern = re.compile('\A(\w*)-fin(\d*)-1Y-REV-[K|N]-HR.xlsx?\Z')
    return pattern.match(filename)


def int_or_zero(s):
    '''
    Ulazni parametar konvertira u integer ukoliko je moguće
    Ukoliko je riječ o praznom stringu, vraća 0
    '''
    s = str(s).strip()
    return int(s) if s else 0


def float_or_zero(s):
    '''
    Ulazni parametar konvertira u float ukoliko je moguće
    Ukoliko je riječ o praznom stringu, vraća 0.0
    '''
    s = str(s).strip()
    return float(s) if s else 0.0


def format_number(s):
    '''
    Ulazni parametar pretvara u float, dobiveni float konvertira u
    integer, te na kraju taj integer konvertira u string i vraća
    kao izlaz
    '''
    return str(int(float_or_zero(s)))


def save_company_data(comp_data, val, additional_data):
    '''
    prikupljene podatke pohranjuje u bazu
    '''
    if len(val) == gfi_expected_items_no:
        company_abbreviation = comp_data[company_abbreviation_key]
        company_name = comp_data[company_name_key]
        company_id = select_or_insert_company_id(company_abbreviation, company_name)
        # create GFI header record
        hdr = GfiHeaders(load_date=datetime.now()
            , original_file_name=additional_data[gfi_orig_filename_key]
            , file_name=additional_data[gfi_dest_filename_key]
            , company = Companies.objects.filter(id=company_id).first()
            , company_name=comp_data[company_name_key]
            , year=additional_data[gfi_year_key])
        hdr.save()
        logging.info("\tHeader {0} created".format(hdr))
        for v in val:
            aop_mark = v[2]
            aop_mark_formatted = str(int(aop_mark)).rjust(3, '0')
            last_year_val_formatted = format_number(v[3]).rjust(12, ' ')
            current_year_val_formatted = format_number(v[4]).rjust(12, ' ')
            report_name = v[0]
            li = select_or_insert_list_item_id(report_name, v[1], aop_mark)
            dtl = GfiDetails(header=hdr
                , item=ListItems.objects.filter(id=li).first()
                , value=current_year_val_formatted
                , value_last_year=last_year_val_formatted)
            dtl.save()
            logging.info("\t\tLine created: {0} {1} {2} {3}".format(str(li).rjust(3, ' '), aop_mark_formatted, last_year_val_formatted, current_year_val_formatted))
        return save_status_ok
    else:
        logging.error("\tUnexpected number of recognized items in file: {no_items}".format(no_items=len(val)))
        return save_status_error

def print_company_data(comp_data, val):
    '''
    ispisuje podatke (radi testiranja)
    '''
    name = company_data[company_name_key]
    abbrev = company_data[company_abbreviation_key]
    print('Kratica: ', abbrev)
    print('Naziv: ', name)
    for v in val:
        # print(v[0][:100].replace('\n', ' ').ljust(100, ' ')
        #     , str(int(v[1])).rjust(3, '0')
        #     , format_number(v[2]).rjust(12, ' ')
        #     , format_number(v[3]).rjust(12, ' '))
        report_name = v[0]
        pos_name = v[1]
        aop_mark = v[2]

        aop_mark_formatted = str(int(v[2])).rjust(3, '0')
        last_year_val_formatted = format_number(v[3]).rjust(12, ' ')
        current_year_val_formatted = format_number(v[4]).rjust(12, ' ')

        li = get_list_item_id(report_name, pos_name, aop_mark)
        print(str(li).rjust(3, ' ')
            , aop_mark_formatted
            , last_year_val_formatted
            , current_year_val_formatted)


def get_list_item_id(report_name, pos_name, aop_mark):
    pos_name_stripped = ''.join(e for e in pos_name if e.isalnum())
    list_item = ListItems.objects.filter(report_name=report_name, aop_mark=aop_mark, stripped_title=pos_name_stripped)
    no_items = len(list_item)
    if no_items >= 1:
        id = list_item[0].id
    else:
        id = -1

    return id


def select_or_insert_list_item_id(report_name, pos_name, aop_mark):
    pos_name_stripped = ''.join(e for e in pos_name if e.isalnum())
    id = get_list_item_id(report_name, pos_name, aop_mark)
    if id == -1:
        li = ListItems(report_name=report_name, aop_mark=aop_mark, title=pos_name, stripped_title=pos_name_stripped)
        li.save()
        id = li.id

    return id


def get_company_id(abbrev):
    company = Companies.objects.filter(abbreviation=abbrev)
    no_items = len(company)
    if no_items >= 1:
        id = company[0].id
    else:
        id = -1

    return id


def select_or_insert_company_id(abbrev, name):
    id = get_company_id(abbrev)
    if id == -1:
        short_name = name[:50]
        comp = Companies(name=name, short_name=short_name, abbreviation=abbrev)
        comp.save()
        id = comp.id

    return id


input_files = get_all_input_files(input_folder, input_pattern)
for filename in input_files:
    values = []
    import_bilanca_ok = False
    import_rdg_ok = False

    logging.info("Script begin {0}".format(datetime.now().strftime('%d.%m.%Y %H:%M:%S.%f')))
    logging.info("Fetching {filename}".format(filename=filename))

    if not check_filename(filename):
        logging.error("\tDoes not meet filename construction condition")
        logging.info("\tMoving to error folder")
        shutil.move(filename, error_folder)
        logging.info("\tFile moved to error folder")
    else:
        logging.info("\tMeets filename construction condition")
        _, file_extension = os.path.splitext(filename)

        if file_extension == '.xls':
            wb = open_xls_as_xlsx(filename)
        else:
            wb = load_workbook(filename=filename, read_only=True)
        # print(filename)

        ws_bilanca = wb['Bilanca']
        ws_rdg = wb['RDG']
        worksheets = [ws_bilanca, ws_rdg]

        company_name = find_company_name(ws_bilanca)
        company_abbreviation, year = find_company_abbreviation_and_gfi_year(filename)
        company_data = {company_name_key: company_name, company_abbreviation_key: company_abbreviation}
        additional_data = {gfi_orig_filename_key: filename, gfi_dest_filename_key: filename, gfi_year_key: year}
        # print('Year ', year)

        for ws in worksheets:
            for row in eligible_rows(ws.rows):
                values.append((ws.title, row[0].value, row[8].value, row[9].value, row[10].value))

        # print_company_data(company_data, values)

        if save_company_data(company_data, values, additional_data) == save_status_ok:
            logging.info("\tSave operation successful")
            logging.info("\tMoving to backup folder")
            shutil.move(filename, error_folder)
            logging.info("\tFile moved to backup folder")
        else:
            logging.error("\tError occured duting save operation")
            logging.info("\tMoving to error folder")
            shutil.move(filename, error_folder)
            logging.info("\tFile moved to error folder")

        print("Import ended")
