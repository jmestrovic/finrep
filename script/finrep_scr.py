from openpyxl import Workbook, load_workbook
import xlrd
import os
import fnmatch
import re

input_folder = 'input'
archive_folder = 'archive'
error_folder = 'error'
input_pattern = '*.xls*'


def get_all_input_files(folder, pattern):
    """
    browse input_folder; search  files whose name matches with input_pattern
    """
    for root, dirs, files in os.walk(folder):
        for basename in files:
            if fnmatch.fnmatch(basename, pattern):
                filename = os.path.join(root, basename)
                yield filename


def open_xls_as_xlsx(fname):
    """
    load xls file with file name fname, convert it to
    xlsx format (in memory) an return converted file
    """
    xlsBook = xlrd.open_workbook(filename=fname)
    workbook = Workbook()
    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

    return workbook


def is_row_empty(row):
    '''
    checks every cell value in array row
    returns True if all values are empty or None
    '''
    is_empty = True
    for cell in row:
        if str(cell.value):
            is_empty = False
            break
    return is_empty


def is_data_row(row):
    '''
    preskoči potpuno prazne retke te one koji u kolonama 1-6 imaju
    neki podatak (takvi reci ne sadrže financijske podatke)
    '''
    is_data_row = (not is_row_empty(row)) \
        and is_row_empty(row[1:7]) \
        and isinstance(row[8].value, (int, float))
    return is_data_row


def eligible_rows(rows):
    '''
    Preskače prvih x redaka na sheetu (zaglavlje)
    Također, preskače i sve ostale retke koji ne sadrže podatke
    '''
    r = []
    is_header = True
    for row in rows:
        if (str(row[0].value).strip().upper().startswith('A)  POTRAŽIVANJA') or(str(row[0].value).strip().startswith('I. POSLOVNI'))):
            is_header = False
        if (not is_header) and is_data_row(row):
            r.append(row)
    return r


def find_company_name(rows):
    '''
    Traži redak koji sadrži naziv tvrtke, te taj naziv ekstrahira i vraća
    '''
    name = ''
    pattern = '\AOBVEZNIK:*\s*(.*)[\s\t]*\Z'
    for row in rows:
        val = (row[0].value).replace('_', ' ').strip().upper()
        if val.startswith('OBVEZNIK'):
            try:
                res = re.search(pattern, val)
                name = res.groups(1)[0]
            except AttributeError:
                pass # not found
            break
    return name


def int_or_zero(s):
    '''
    Ulazni parametar konvertira u integer ukoliko je moguće
    Ukoliko je riječ o praznom stringu, vraća 0
    '''
    s = str(s).strip()
    return int(s) if s else 0


def float_or_zero(s):
    '''
    Ulazni parametar konvertira u float ukoliko je moguće
    Ukoliko je riječ o praznom stringu, vraća 0.0
    '''
    s = str(s).strip()
    return float(s) if s else 0.0


def format_number(s):
    '''
    Ulazni parametar pretvara u float, dobiveni float konvertira u
    integer, te na kraju taj integer konvertira u string i vraća 
    kao izlaz
    '''
    return str(int(float_or_zero(s)))


def print_company_data(name, val):
    print(company_name)
    for v in val:
        print(v[0][:100].replace('\n', ' ').ljust(100, ' ')
            , str(int(v[1])).rjust(3, '0')
            , format_number(v[2]).rjust(12, ' ')
            , format_number(v[3]).rjust(12, ' '))
        pass
        try:
            pass
            #print(v[0].value[:100].ljust(100, ' ')
            #    , str(int(v[1].value)).rjust(3, '0')
            #    , str(int(v[2].value)).rjust(15, ' ')
            #    , str(int(v[3].value)).rjust(15, ' '))
        except:
            pass


input_files = get_all_input_files(input_folder, input_pattern)
for filename in input_files:
    values = []
    import_bilanca_ok = False
    import_rdg_ok = False
    _, file_extension = os.path.splitext(filename)

    if file_extension == '.xls':
        wb = open_xls_as_xlsx(filename)
    else:
        wb = load_workbook(filename=filename, read_only=True)
    #print(filename)

    ws_bilanca = wb['Bilanca']
    ws_rdg = wb['RDG']
    worksheets = [ws_bilanca, ws_rdg]

    company_name = find_company_name(ws_bilanca)

    for ws in worksheets:
        for row in eligible_rows(ws.rows):
            #print((row[0].value, row[8].value, row[9].value, row[10].value))
            values.append((row[0].value, row[8].value, row[9].value, row[10].value))


    print_company_data(company_name, values)
